#!/usr/bin/env python3
"""
Excel 数据读取工具 — 读取 xlsx 并输出纯文本表格

用法:
    python3 tools/read-excel.py 数据.xlsx
依赖:
    pip install openpyxl
"""

import sys
from pathlib import Path
try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ 请先安装: pip install openpyxl")
    sys.exit(1)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"❌ 文件不存在: {path}")
        sys.exit(1)

    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows()):
        vals = [
            str(c.value or "").replace("\n", "↵").strip() for c in row
        ]
        print(" [COL] ".join(vals))
        if i < ws.max_row - 1:
            print("---")


if __name__ == "__main__":
    main()
